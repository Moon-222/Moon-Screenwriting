import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

EXCEL_EXT = {".xlsx", ".xlsm"}
SKIP_PARTS = {".git", ".github", "node_modules", "__pycache__", ".venv", "venv"}


def iter_excel_files(root: Path):
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() not in EXCEL_EXT:
            continue
        rel = path.relative_to(root)
        if any(part in SKIP_PARTS for part in rel.parts):
            continue
        yield path


def sanitize_sheet_name(name: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in str(name))
    return cleaned or "sheet"


def convert(excel_path: Path) -> list[Path]:
    written: list[Path] = []
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            stem = excel_path.stem
            if len(sheet_names) == 1:
                csv_path = excel_path.with_suffix(".csv")
            else:
                csv_path = excel_path.with_name(f"{stem}_{sanitize_sheet_name(sheet_name)}.csv")
            with csv_path.open("w", newline="", encoding="utf-8-sig") as fh:
                writer = csv.writer(fh)
                for row in ws.iter_rows(values_only=True):
                    if row is None:
                        continue
                    if all(cell is None or (isinstance(cell, str) and cell == "") for cell in row):
                        continue
                    writer.writerow(["" if cell is None else cell for cell in row])
            print(f"converted: {excel_path} [{sheet_name}] -> {csv_path}")
            written.append(csv_path)
    finally:
        wb.close()
    return written


def main() -> int:
    root = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(".")
    if not root.is_dir():
        print(f"root not found: {root}", file=sys.stderr)
        return 1
    total = 0
    for excel_path in iter_excel_files(root):
        try:
            total += len(convert(excel_path))
        except Exception as exc:
            print(f"failed: {excel_path}: {exc}", file=sys.stderr)
            return 1
    print(f"done: {total} csv file(s) written")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
